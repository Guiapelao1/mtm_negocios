import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook

#calcular a quantidade total de insumos necessários
def calcular_quantidade_insumos(insumos_por_produto, producao):
    quantidade_insumos = {'Insumo 1': 0, 'Insumo 2': 0, 'Insumo 3': 0}
    
    for produto, quantidade in producao.items():
        for insumo, valor in insumos_por_produto[produto].items():
            quantidade_insumos[insumo] += valor * quantidade
    
    return quantidade_insumos

#calcular o custo total de produção
def calcular_custo_total(quantidade_insumos, custos):
    custo_total = 0
    for insumo, quantidade in quantidade_insumos.items():
        custo_total += quantidade * custos[insumo]
    return custo_total

def exportar_para_excel(quantidade_insumos, custo_total_atual, custo_total_aumentado, impacto, custos_aumentados):
    # Cria um novo arquivo Excel
    wb = Workbook()
    ws = wb.active

    # Adiciona os cabeçalhos
    ws['A1'] = "Item"
    ws['B1'] = "Valor"

    # Adiciona os dados
    ws['A2'] = "Quantidade de Insumo 1 (kg)"
    ws['B2'] = quantidade_insumos['Insumo 1']
    ws['A3'] = "Quantidade de Insumo 2 (kg)"
    ws['B3'] = quantidade_insumos['Insumo 2']
    ws['A4'] = "Quantidade de Insumo 3 (kg)"
    ws['B4'] = quantidade_insumos['Insumo 3']
    ws['A5'] = "Custo Total Atual (R$)"
    ws['B5'] = custo_total_atual
    ws['A6'] = "Custo Total com Aumento (R$)"
    ws['B6'] = custo_total_aumentado
    ws['A7'] = "Impacto do Aumento (R$)"
    ws['B7'] = impacto

    # Adiciona os valores dos aumentos
    ws['A9'] = "Aumentos Aplicados"
    ws['A10'] = "Insumo"
    ws['B10'] = "Valor do Aumento por kg (R$/kg)"
    ws['C10'] = "Valor Total do Aumento (R$)"
    linha = 11
    valor_total_aumentos = 0  # Variável para acumular o valor total dos aumentos
    for insumo, aumento in custos_aumentados.items():
        if aumento > 0:
            valor_total_insumo = aumento * quantidade_insumos[insumo]  # Valor total do aumento para o insumo
            valor_total_aumentos += valor_total_insumo  # Acumula o valor total dos aumentos
            ws[f'A{linha}'] = insumo
            ws[f'B{linha}'] = aumento  # Valor do aumento por kg
            ws[f'C{linha}'] = valor_total_insumo  # Valor total do aumento para o insumo
            linha += 1

    # Adiciona o valor total dos aumentos
    ws[f'A{linha}'] = "Valor Total dos Aumentos"
    ws[f'C{linha}'] = valor_total_aumentos

    # Salva o arquivo
    wb.save("resultados.xlsx")
    messagebox.showinfo("Exportado", "Os resultados foram exportados para 'resultados.xlsx'.")

# Função principal que será chamada ao clicar no botão "Calcular"
def calcular():
    try:
        # Coletando os dados dos campos de entrada
        insumos_por_produto = {
            'Produto A': {'Insumo 1': float(entrada_insumo1_A.get()), 'Insumo 2': float(entrada_insumo2_A.get()), 'Insumo 3': float(entrada_insumo3_A.get())},
            'Produto B': {'Insumo 1': float(entrada_insumo1_B.get()), 'Insumo 2': float(entrada_insumo2_B.get()), 'Insumo 3': float(entrada_insumo3_B.get())},
            'Produto C': {'Insumo 1': float(entrada_insumo1_C.get()), 'Insumo 2': float(entrada_insumo2_C.get()), 'Insumo 3': float(entrada_insumo3_C.get())}
        }

        producao = {
            'Produto A': float(entrada_producao_A.get()),
            'Produto B': float(entrada_producao_B.get()),
            'Produto C': float(entrada_producao_C.get())
        }

        custos_atuais = {
            'Insumo 1': float(entrada_custo_insumo1.get()),
            'Insumo 2': float(entrada_custo_insumo2.get()),
            'Insumo 3': float(entrada_custo_insumo3.get())
        }

        # Coletando os aumentos dos insumos
        custos_aumentados = {
            'Insumo 1': float(entrada_aumento_insumo1.get()),
            'Insumo 2': float(entrada_aumento_insumo2.get()),
            'Insumo 3': float(entrada_aumento_insumo3.get())
        }

        # Calculando a quantidade total de insumos necessários
        quantidade_insumos = calcular_quantidade_insumos(insumos_por_produto, producao)

        # Calculando o custo total de produção com os custos atuais
        custo_total_atual = calcular_custo_total(quantidade_insumos, custos_atuais)

        # Aplicando os aumentos nos custos
        custos_com_aumento = custos_atuais.copy()
        for insumo, aumento in custos_aumentados.items():
            if aumento > 0:
                custos_com_aumento[insumo] += aumento

        # Calculando o custo total de produção com os aumentos
        custo_total_aumentado = calcular_custo_total(quantidade_insumos, custos_com_aumento)

        # Calculando o impacto dos aumentos
        impacto = custo_total_aumentado - custo_total_atual

        # Exibindo os resultados em uma caixa de mensagem
        resultado = (
            f"Quantidade total de insumos necessários:\n"
            f"Insumo 1: {quantidade_insumos['Insumo 1']} kg\n"
            f"Insumo 2: {quantidade_insumos['Insumo 2']} kg\n"
            f"Insumo 3: {quantidade_insumos['Insumo 3']} kg\n\n"
            f"Custo total de produção com os custos atuais: R$ {custo_total_atual:.2f}\n"
            f"Custo total de produção com os aumentos: R$ {custo_total_aumentado:.2f}\n\n"
            f"Impacto dos aumentos: R$ {impacto:.2f}"
        )
        messagebox.showinfo("Resultados", resultado)

        # Exporta os resultados para o Excel
        exportar_para_excel(quantidade_insumos, custo_total_atual, custo_total_aumentado, impacto, custos_aumentados)

    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira valores numéricos em todos os campos.")

# Criando a janela principal
janela = tk.Tk()
janela.title("Cálculo de Custos de Produção")

# Criando e posicionando os widgets na janela
tk.Label(janela, text="Insumos por Produto (kg/unidade)").grid(row=0, column=0, columnspan=4, pady=10)

# Produto A
tk.Label(janela, text="Produto A").grid(row=1, column=0)
tk.Label(janela, text="Insumo 1:").grid(row=2, column=0)
entrada_insumo1_A = tk.Entry(janela)
entrada_insumo1_A.grid(row=2, column=1)
tk.Label(janela, text="Insumo 2:").grid(row=3, column=0)
entrada_insumo2_A = tk.Entry(janela)
entrada_insumo2_A.grid(row=3, column=1)
tk.Label(janela, text="Insumo 3:").grid(row=4, column=0)
entrada_insumo3_A = tk.Entry(janela)
entrada_insumo3_A.grid(row=4, column=1)

# Produto B
tk.Label(janela, text="Produto B").grid(row=1, column=2)
tk.Label(janela, text="Insumo 1:").grid(row=2, column=2)
entrada_insumo1_B = tk.Entry(janela)
entrada_insumo1_B.grid(row=2, column=3)
tk.Label(janela, text="Insumo 2:").grid(row=3, column=2)
entrada_insumo2_B = tk.Entry(janela)
entrada_insumo2_B.grid(row=3, column=3)
tk.Label(janela, text="Insumo 3:").grid(row=4, column=2)
entrada_insumo3_B = tk.Entry(janela)
entrada_insumo3_B.grid(row=4, column=3)

# Produto C
tk.Label(janela, text="Produto C").grid(row=1, column=4)
tk.Label(janela, text="Insumo 1:").grid(row=2, column=4)
entrada_insumo1_C = tk.Entry(janela)
entrada_insumo1_C.grid(row=2, column=5)
tk.Label(janela, text="Insumo 2:").grid(row=3, column=4)
entrada_insumo2_C = tk.Entry(janela)
entrada_insumo2_C.grid(row=3, column=5)
tk.Label(janela, text="Insumo 3:").grid(row=4, column=4)
entrada_insumo3_C = tk.Entry(janela)
entrada_insumo3_C.grid(row=4, column=5)

# Previsão de Produção
tk.Label(janela, text="Previsão de Produção (unidades)").grid(row=5, column=0, columnspan=6, pady=10)
tk.Label(janela, text="Produto A:").grid(row=6, column=0)
entrada_producao_A = tk.Entry(janela)
entrada_producao_A.grid(row=6, column=1)
tk.Label(janela, text="Produto B:").grid(row=6, column=2)
entrada_producao_B = tk.Entry(janela)
entrada_producao_B.grid(row=6, column=3)
tk.Label(janela, text="Produto C:").grid(row=6, column=4)
entrada_producao_C = tk.Entry(janela)
entrada_producao_C.grid(row=6, column=5)

# Custos dos Insumos
tk.Label(janela, text="Custos dos Insumos (R$/kg)").grid(row=7, column=0, columnspan=6, pady=10)
tk.Label(janela, text="Insumo 1:").grid(row=8, column=0)
entrada_custo_insumo1 = tk.Entry(janela)
entrada_custo_insumo1.grid(row=8, column=1)
tk.Label(janela, text="Insumo 2:").grid(row=8, column=2)
entrada_custo_insumo2 = tk.Entry(janela)
entrada_custo_insumo2.grid(row=8, column=3)
tk.Label(janela, text="Insumo 3:").grid(row=8, column=4)
entrada_custo_insumo3 = tk.Entry(janela)
entrada_custo_insumo3.grid(row=8, column=5)

# Aumentos dos Insumos
tk.Label(janela, text="Aumentos dos Insumos (R$/kg)").grid(row=9, column=0, columnspan=6, pady=10)
tk.Label(janela, text="Insumo 1:").grid(row=10, column=0)
entrada_aumento_insumo1 = tk.Entry(janela)
entrada_aumento_insumo1.grid(row=10, column=1)
tk.Label(janela, text="Insumo 2:").grid(row=10, column=2)
entrada_aumento_insumo2 = tk.Entry(janela)
entrada_aumento_insumo2.grid(row=10, column=3)
tk.Label(janela, text="Insumo 3:").grid(row=10, column=4)
entrada_aumento_insumo3 = tk.Entry(janela)
entrada_aumento_insumo3.grid(row=10, column=5)

# Botão para calcular
botao_calcular = tk.Button(janela, text="Calcular", command=calcular)
botao_calcular.grid(row=11, column=0, columnspan=6, pady=20)

# Iniciando a interface gráfica
janela.mainloop()